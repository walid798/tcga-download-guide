import os
import json
import shutil
from pathlib import Path
from openpyxl import Workbook, load_workbook

# -----------------------------
# Config
# -----------------------------
RAW_DIR = Path(r"slides\raw")
ORG_DIR = Path(r"slides\organized")
EXCEL_PATH = Path(r"slides\cases.xlsx")

# Where your JSON metadata lives
METADATA_DIR = Path(r"Kidney\Metadata")

# Which labels to process
LABELS = ["KIRC", "KIRP", "KICH"]

# Which file extensions to treat as slide files
SLIDE_EXTS = {".svs"}
# , ".tif", ".tiff", ".ndpi", ".mrxs"
# If True: do not move files, only print planned actions
DRY_RUN = False


# -----------------------------
# Helpers
# -----------------------------
def case_id_from_entity_submitter_id(entity_submitter_id: str) -> str:
    """
    Slide submitter IDs often look like: TCGA-XX-YYYY-01Z-00-DX1
    Case ID should be: TCGA-XX-YYYY
    """
    parts = entity_submitter_id.split("-")
    return "-".join(parts[:3]) if len(parts) >= 3 else entity_submitter_id


def safe_move(src: Path, dst: Path) -> Path:
    """
    Move src -> dst. If dst exists, append _1, _2, ... before extension.
    Returns the final destination path.
    """
    dst.parent.mkdir(parents=True, exist_ok=True)

    if not dst.exists():
        if not DRY_RUN:
            shutil.move(str(src), str(dst))
        return dst

    stem, suffix = dst.stem, dst.suffix
    i = 1
    while True:
        candidate = dst.with_name(f"{stem}_{i}{suffix}")
        if not candidate.exists():
            if not DRY_RUN:
                shutil.move(str(src), str(candidate))
            return candidate
        i += 1


def load_metadata_lookup(labels):
    """
    Returns:
      lookup[file_id] = {"label": <label>, "case_id": <case_id>, "file_name": <expected file_name or None>}
    """
    lookup = {}

    for label in labels:
        json_path = METADATA_DIR / f"Metadata_TCGA-{label}.json"
        if not json_path.exists():
            print(f"[WARN] Metadata not found: {json_path}")
            continue

        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        for item in data:
            file_id = item.get("file_id")
            file_name = item.get("file_name")

            assoc = item.get("associated_entities") or []
            entity_submitter_id = None
            if len(assoc) > 0:
                entity_submitter_id = assoc[0].get("entity_submitter_id")

            if not file_id or not entity_submitter_id:
                continue

            case_id = case_id_from_entity_submitter_id(entity_submitter_id)

            # If same file_id appears twice, keep the first occurrence (or overwrite—your choice)
            if file_id not in lookup:
                lookup[file_id] = {
                    "label": label,
                    "case_id": case_id,
                    "file_name": file_name
                }

    return lookup


def load_or_create_excel(path: Path):
    """
    Creates cases.xlsx if missing; otherwise loads it and returns (wb, ws, existing_case_ids_set)
    """
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
        existing = set()

        # Expect header: Case_id, Label
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                existing.add(str(row[0]).strip())
        return wb, ws, existing

    wb = Workbook()
    ws = wb.active
    ws.title = "Cases"
    ws.append(["Case_id", "Label"])
    return wb, ws, set()


# -----------------------------
# Main
# -----------------------------
def main():
    ORG_DIR.mkdir(parents=True, exist_ok=True)

    # 1) Build file_id -> case_id/label map from all JSONs
    file_id_lookup = load_metadata_lookup(LABELS)
    print(f"[INFO] Loaded {len(file_id_lookup)} file_id entries from metadata.")

    # 2) Load shared excel (append new cases only)
    wb, ws, existing_cases = load_or_create_excel(EXCEL_PATH)

    moved_files = 0
    newly_logged_cases = 0

    # 3) Iterate raw folders (each folder is expected to be a file_id)
    for folder_name in os.listdir(RAW_DIR):
        file_id_dir = RAW_DIR / folder_name
        if not file_id_dir.is_dir():
            continue

        file_id = folder_name
        meta = file_id_lookup.get(file_id)

        if not meta:
            print(f"[SKIP] {file_id} not found in metadata.")
            continue

        label = meta["label"]
        case_id = meta["case_id"]
        expected_name = meta.get("file_name")

        case_dir = ORG_DIR / case_id
        case_dir.mkdir(parents=True, exist_ok=True)

        # Collect candidate slide files in this file_id directory
        files_in_dir = [p for p in file_id_dir.iterdir() if p.is_file()]
        slide_files = [p for p in files_in_dir if p.suffix.lower() in SLIDE_EXTS]

        # Prefer moving the expected filename if present; otherwise move all slide files found
        to_move = []
        if expected_name:
            expected_path = file_id_dir / expected_name
            if expected_path.exists():
                to_move = [expected_path]
            else:
                # fallback: move all slide files
                to_move = slide_files
                print(f"[WARN] Expected file not found for {file_id}: {expected_name}. Falling back to slide files in folder.")
        else:
            to_move = slide_files

        if not to_move:
            print(f"[WARN] No slide files found to move in: {file_id_dir}")
            continue

        # Move files
        for src in to_move:
            dst = case_dir / src.name
            final_dst = safe_move(src, dst)
            moved_files += 1
            print(f"[MOVE] {src.name} -> {final_dst}")

        # Remove empty file_id folder if everything moved (optional)
        remaining = list(file_id_dir.iterdir())
        if len(remaining) == 0:
            if not DRY_RUN:
                file_id_dir.rmdir()
            print(f"[CLEAN] Removed empty folder: {file_id_dir}")

        # Log case once in Excel
        if case_id not in existing_cases:
            ws.append([case_id, label])
            existing_cases.add(case_id)
            newly_logged_cases += 1

    # 4) Save Excel
    if not DRY_RUN:
        wb.save(EXCEL_PATH)

    print("\n----- Summary -----")
    print(f"Moved files: {moved_files}")
    print(f"New cases added to Excel: {newly_logged_cases}")
    print(f"Excel saved at: {EXCEL_PATH}")
    if DRY_RUN:
        print("[DRY_RUN] No files were actually moved.")


if __name__ == "__main__":
    main()
